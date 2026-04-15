#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Web应用 - 提供Excel上传页面并调用API处理
支持AI回答和语义对比打分
"""

import os
import re
import json
import ssl
import logging
import urllib.request
import queue
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 最大16MB
app.config['JSON_AS_ASCII'] = False

# 确保文件夹存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 评分模型配置 - 使用内网API
SCORING_API_URL = os.environ.get('ANTHROPIC_BASE_URL', 'http://ai.tech.tax.asia.pwcinternal.com:3002') + '/v1/chat/completions'
SCORING_API_KEY = os.environ.get('ANTHROPIC_AUTH_TOKEN', '')
SCORING_MODEL = os.environ.get('ANTHROPIC_MODEL', 'glm-coding-5-8')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def request_api(message: str, session_id: str = "") -> tuple:
    """请求知识库API接口"""
    url = 'https://ai.tech.tax.asia.pwcinternal.com:5007/api/chat-stream'
    data = json.dumps({
        'message': message,
        'session_id': session_id,
        'model': 'GLM-5.1'
    }).encode('utf-8')

    req = urllib.request.Request(url, data=data, headers={
        'Content-Type': 'application/json',
        'Accept': 'text/event-stream'
    })

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    with urllib.request.urlopen(req, context=ctx, timeout=300) as response:
        returned_session_id = response.headers.get('X-Session-Id', '')
        result = response.read().decode('utf-8')

    return result, returned_session_id


def parse_response(api_response: str) -> dict:
    """解析SSE流式API响应"""
    contents = []
    for line in api_response.strip().split('\n'):
        line = line.strip()
        if line.startswith('data:'):
            try:
                data = json.loads(line[5:].strip())
                if data.get('type') == 'content':
                    contents.append(data.get('content', ''))
            except json.JSONDecodeError:
                pass

    full = ''.join(contents)

    # 截取"核心发现"之后的内容，去掉前面的工具调用思考过程
    marker = '核心发现'
    idx = full.find(marker)
    if idx != -1:
        full = full[idx:]

    return {'full_content': full}


def is_confirmation_question(content: str) -> bool:
    """判断内容是否是确认问句"""
    keywords = [
        '请问以上关键词是否需要调整或补充', '确认后我将开始',
        '请问这样理解是否正确', '您是否有需要补充或调整',
        '是否需要调整', '请确认', '确认后', '以上关键词是否准确',
        '是否需要添加', '请问以上', '确认后开始', '是否准确',
        '请告知', '是否继续', '我将开始', '以上内容是否',
        '是否合适', '希望调整', '是否需要修改', '是否满意',
        '是否同意', '请告诉我您的修改', '需要调整请告诉我',
    ]
    return any(kw in content for kw in keywords)


def is_incomplete_answer(content: str) -> bool:
    """判断内容是否是不完整的中间步骤（没有核心发现且较短）"""
    if '核心发现' in content:
        return False
    # 内容过短说明还没拿到完整答案
    return len(content.strip()) < 200


def chat_with_confirmation(question: str, max_rounds: int = 8) -> str:
    """执行多轮对话，自动处理确认和等待完整答案"""
    session_id = ""
    current_message = question

    for _ in range(max_rounds):
        api_response, session_id = request_api(current_message, session_id)
        content = parse_response(api_response)['full_content']
        if is_confirmation_question(content):
            current_message = "同意，请使用这些关键词进行搜索，不需要调整。"
        elif is_incomplete_answer(content):
            # 拿到的是中间步骤，继续等待完整答案
            current_message = "继续"
        else:
            return content

    return content


def request_scoring_api(prompt: str) -> str:
    """请求内网AI评分API"""
    data = json.dumps({
        'model': SCORING_MODEL,
        'messages': [{'role': 'user', 'content': prompt}]
    }).encode('utf-8')

    req = urllib.request.Request(SCORING_API_URL, data=data, headers={
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {SCORING_API_KEY}'
    })

    with urllib.request.urlopen(req, timeout=180) as response:
        result = json.loads(response.read().decode('utf-8'))
        if 'choices' not in result or not result['choices']:
            raise ValueError(f"API返回格式异常: {json.dumps(result, ensure_ascii=False)[:300]}")
        return result['choices'][0]['message']['content']


def score_answer(question: str, ai_answer: str, reference_answer: str) -> dict:
    """使用AI对回答进行语义对比打分"""
    scoring_prompt = f"""请作为一名专业的税务领域评估专家，对以下AI回答进行评分。

【问题】
{question}

【参考答案】
{reference_answer}

【AI回答】
{ai_answer}

请从以下5个维度进行评分（每个维度0-20分，满分20分，如果该维度表现优秀应当给满分，不要为了避免给满分而刻意扣分），并以JSON格式返回结果：

1. 答案准确性：回答内容是否准确，与参考答案的核心观点是否一致。完全一致给20分
2. 法条援引度：是否正确引用了相关法规条款。援引完整正确给20分
3. 逻辑清晰度：回答的逻辑结构是否清晰，论证是否严密。逻辑严密给20分
4. 格式化表示：回答的格式是否清晰易读，是否有良好的结构化呈现。格式优秀给20分
5. 总结完整度：总结是否完整，是否涵盖了问题的各个方面。总结全面给20分

请严格按照以下JSON格式返回（不要添加任何其他文字）：
{{
    "accuracy_score": 分数,
    "accuracy_reason": "评分说明",
    "citation_score": 分数,
    "citation_reason": "评分说明",
    "logic_score": 分数,
    "logic_reason": "评分说明",
    "format_score": 分数,
    "format_reason": "评分说明",
    "summary_score": 分数,
    "summary_reason": "评分说明"
}}
"""

    failed = {
        'success': False,
        'accuracy_score': 0, 'accuracy_reason': '评分失败',
        'citation_score': 0, 'citation_reason': '评分失败',
        'logic_score': 0, 'logic_reason': '评分失败',
        'format_score': 0, 'format_reason': '评分失败',
        'summary_score': 0, 'summary_reason': '评分失败',
        'total_score': 0
    }

    last_error = None
    for attempt in range(3):
        try:
            content = request_scoring_api(scoring_prompt)

            json_start = content.find('{')
            json_end = content.rfind('}') + 1
            if json_start == -1 or json_end <= json_start:
                last_error = f"评分返回无JSON: {content[:200]}"
                logger.error(f"第{attempt+1}次尝试 - {last_error}")
                continue

            scores = json.loads(content[json_start:json_end])
            score_keys = ['accuracy_score', 'citation_score', 'logic_score', 'format_score', 'summary_score']
            return {
                'success': True,
                **{k: scores.get(k, 0) for k in score_keys},
                **{k.replace('score', 'reason'): scores.get(k.replace('score', 'reason'), '') for k in score_keys},
                'total_score': sum(scores.get(k, 0) for k in score_keys)
            }
        except Exception as e:
            last_error = str(e)
            logger.error(f"第{attempt+1}次评分失败: {e}")
            if attempt < 2:
                import time
                time.sleep(2)

    # 3次都失败，把具体错误写入reason
    failed['accuracy_reason'] = f'评分失败({last_error})' if last_error else '评分失败'
    failed['citation_reason'] = failed['accuracy_reason']
    failed['logic_reason'] = failed['accuracy_reason']
    failed['format_reason'] = failed['accuracy_reason']
    failed['summary_reason'] = failed['accuracy_reason']
    return failed


def read_questions_from_excel(filepath):
    """从Excel文件读取问题(B列)和建议答案(E列)"""
    wb = load_workbook(filepath)
    ws = wb.active
    questions = []
    row = 2

    while True:
        cell_value = ws[f'B{row}'].value
        if cell_value is None or str(cell_value).strip() == '':
            break

        reference_answer = ws[f'E{row}'].value
        reference_answer = str(reference_answer).strip() if reference_answer else ''
        questions.append((row, str(cell_value).strip(), reference_answer))
        row += 1

    return questions


def save_results_to_excel(questions_with_answers, original_filepath, output_filepath):
    """保存结果到Excel，保留原格式"""
    wb = load_workbook(original_filepath)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    max_col = ws.max_column
    start_col = max_col + 1

    headers = [
        ('AI回答', 100),
        ('答案准确性(20分)', 12), ('答案准确性说明', 40),
        ('法条援引度(20分)', 12), ('法条援引度说明', 40),
        ('逻辑清晰度(20分)', 12), ('逻辑清晰度说明', 40),
        ('格式化表示(20分)', 12), ('格式化表示说明', 40),
        ('总结完整度(20分)', 12), ('总结完整度说明', 40),
        ('总分(100分)', 10)
    ]

    for i, (header, width) in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    for row_num, question, answer, scores in questions_with_answers:
        cell = ws.cell(row=row_num, column=start_col)
        cell.value = answer
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

        if scores:
            score_data = [
                scores.get('accuracy_score', 0),
                scores.get('accuracy_reason', ''),
                scores.get('citation_score', 0),
                scores.get('citation_reason', ''),
                scores.get('logic_score', 0),
                scores.get('logic_reason', ''),
                scores.get('format_score', 0),
                scores.get('format_reason', ''),
                scores.get('summary_score', 0),
                scores.get('summary_reason', ''),
                scores.get('total_score', 0)
            ]
            for i, value in enumerate(score_data):
                cell = ws.cell(row=row_num, column=start_col + 1 + i)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border

    wb.save(output_filepath)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not (file and allowed_file(file.filename)):
        return jsonify({'error': '不支持的文件格式'}), 400

    # 保留中文文件名，只去掉不安全字符
    safe_name = re.sub(r'[\\/*?:"<>|]', '_', file.filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{safe_name}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        questions = read_questions_from_excel(filepath)
    except Exception as e:
        return jsonify({'error': f'读取Excel文件失败: {str(e)}'}), 400

    return jsonify({
        'success': True,
        'filename': filename,
        'question_count': len(questions),
        'questions': [{'row': r, 'question': q, 'has_reference': bool(ref)} for r, q, ref in questions]
    })


def process_single_question(row_num, question, reference_answer, enable_scoring):
    """处理单个问题：获取AI回答 + 评分"""
    try:
        answer = chat_with_confirmation(question)

        scores = None
        if enable_scoring and reference_answer:
            scores = score_answer(question, answer, reference_answer)

        return {
            'row': row_num,
            'question': question,
            'answer': answer,
            'reference_answer': reference_answer,
            'scores': scores,
            'success': True
        }
    except Exception as e:
        logger.error(f"问题处理失败 (行{row_num}): {e}")
        return {
            'row': row_num,
            'question': question,
            'answer': f'处理失败: {str(e)}',
            'reference_answer': reference_answer,
            'scores': None,
            'success': False
        }


@app.route('/process', methods=['POST'])
def process_questions():
    """SSE流式处理接口，多线程并发，实时返回每题进度"""
    data = request.json
    filename = data.get('filename')
    enable_scoring = data.get('enable_scoring', False)
    thread_count = data.get('thread_count', 2)
    thread_count = max(1, min(8, int(thread_count)))

    def sse_error(msg):
        def gen():
            yield f"data: {json.dumps({'type': 'error', 'message': msg}, ensure_ascii=False)}\n\n"
        return app.response_class(gen(), mimetype='text/event-stream',
                                  headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

    if not filename:
        return sse_error('缺少文件名')

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return sse_error('文件不存在')

    questions = read_questions_from_excel(filepath)
    total = len(questions)
    result_queue = queue.Queue()

    def worker(row_num, question, reference_answer):
        result = process_single_question(row_num, question, reference_answer, enable_scoring)
        result_queue.put(result)

    def generate():
        # 提交所有任务到线程池
        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            futures = []
            for row_num, question, reference_answer in questions:
                futures.append(executor.submit(worker, row_num, question, reference_answer))

            # 实时收集已完成的结果
            completed = 0
            results = []
            while completed < total:
                result = result_queue.get()
                results.append(result)
                completed += 1

                event = {
                    'type': 'progress',
                    'current': completed,
                    'total': total,
                    'percentage': int(completed / total * 100),
                    'result': result
                }
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

        # 按行号排序后保存结果
        results.sort(key=lambda r: r['row'])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"AI回答结果_{timestamp}.xlsx"
        output_filepath = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        questions_with_answers = [(r['row'], r['question'], r['answer'], r['scores']) for r in results]
        save_results_to_excel(questions_with_answers, filepath, output_filepath)

        yield f"data: {json.dumps({'type': 'complete', 'output_filename': output_filename}, ensure_ascii=False)}\n\n"

    return app.response_class(generate(), mimetype='text/event-stream',
                              headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': '文件不存在'}), 404


if __name__ == '__main__':
    app.run(debug=True, port=5001, host='0.0.0.0')
